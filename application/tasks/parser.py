from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class Parser:
    def __init__(self):
        self.data_in: Worksheet = load_workbook(
            filename="/Users/gglol/PycharmProjects/Y_lab_fastapi_Test/application/admin/Menu.xlsx"
        ).active
        self.data_out = []

    def get_menu_data(self, row: int) -> dict | None:
        result = {}
        menu_id = self.data_in[f"A{row}"].value
        menu_title, menu_description = (
            self.data_in[f"B{row}"].value,
            self.data_in[f"C{row}"].value,
        )
        data = {
            "id": menu_id,
            "title": menu_title,
            "description": menu_description,
            "submenus": [],
        }
        result.update(data)
        for i in range(row + 1, self.data_in.max_row + 1):
            menu_id = self.data_in[f"A{i}"].value
            if not menu_id:
                submenu_id = self.data_in[f"B{i}"].value
                if submenu_id:
                    result["submenus"].append(self.get_submenu_data(i))
        return result

    def get_submenu_data(self, row: int) -> dict | None:
        result = {}
        submenu_id = self.data_in[f"B{row}"].value
        submenu_title, submenu_description = (
            self.data_in[f"C{row}"].value,
            self.data_in[f"D{row}"].value,
        )
        data = {
            "id": submenu_id,
            "title": submenu_title,
            "description": submenu_description,
            "dishes": [],
        }
        result.update(data)
        for i in range(row + 1, self.data_in.max_row + 1):
            menu_id = self.data_in[f"A{i}"].value
            submenu_id = self.data_in[f"B{i}"].value
            if not menu_id and not submenu_id:
                result["dishes"].append(self.get_dish_data(i))
            else:
                break
        return result

    def get_dish_data(self, row: int) -> dict | None:
        dish_id = self.data_in[f"C{row}"].value
        if dish_id:
            dish_title, dish_description, dish_price = (
                self.data_in[f"D{row}"].value,
                self.data_in[f"E{row}"].value,
                self.data_in[f"F{row}"].value,
            )
            dish_discount = self.data_in[f"G{row}"].value or 0
            data = {
                "id": dish_id,
                "title": dish_title,
                "description": dish_description,
                "price": dish_price,
                "discount": dish_discount,
            }
            return data

    def start_parser(self):
        self.data_out = []
        for i in range(self.data_in.min_row, self.data_in.max_row + 1):
            menu_id = self.data_in[f"A{i}"].value
            if menu_id:
                self.data_out.append(self.get_menu_data(row=i))
            continue
        return self.data_out
